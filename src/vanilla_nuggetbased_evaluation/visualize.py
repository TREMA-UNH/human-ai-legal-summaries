import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tabulate import tabulate
import numpy as np
from termcolor import colored
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Load JSON file
with open('/Users/nfarzi/Documents/nextpoint/results/evaluation/Anenen_Steve.json', 'r') as file:
    data = json.load(file)

# Extract data for visualizations
consolidated_averages = data['FDC']['consolidated_averages']
fdc_nuggets = data['FDC']['nuggets']
ctc_nuggets = data['CTC']['c_nuggets']

# Prepare data for Bar Chart (Consolidated Averages)
avg_df = pd.DataFrame.from_dict(consolidated_averages, orient='index', columns=['Average Score'])
avg_df.index.name = 'Consolidated ID'

# Prepare data for Stacked Bar Chart (Score Distribution)
score_counts = {cid: {'Score 0': 0, 'Score 1': 0, 'Score 2': 0} for cid in consolidated_averages.keys()}
for nugget in fdc_nuggets:
    cid = nugget['consolidated_id']
    score = nugget['score']
    score_counts[cid][f'Score {score}'] += 1

score_df = pd.DataFrame(score_counts).T
score_df = score_df[['Score 0', 'Score 1', 'Score 2']]  # Ensure consistent order

# Prepare data for Table
table_data = []
for fdc_nugget in fdc_nuggets:
    nugget_id = fdc_nugget['nugget_id']
    cid = fdc_nugget['consolidated_id']
    text = fdc_nugget['text']  # Full text for Excel
    score = fdc_nugget['score']
    explanation = fdc_nugget['explanation']  # Full text for Excel
    
    # Find corresponding CTC nugget for presence and text
    ctc_nugget = next((n for n in ctc_nuggets if n['consolidated_id'] == cid), None)
    presence = ctc_nugget['present'] if ctc_nugget else 'N/A'
    ctc_text = ctc_nugget['text'] if ctc_nugget else 'N/A'
    
    table_data.append([nugget_id, cid, ctc_text, text, presence, score, explanation])

table_df = pd.DataFrame(table_data, columns=['Nugget ID', 'Consolidated ID', 'Consolidated Nugget Text', 'Text', 'Presence', 'Score', 'Explanation'])

# Save table to Excel with conditional formatting
wb = Workbook()
ws = wb.active
ws.title = "Nugget Evaluation"

# Write headers
headers = table_df.columns.tolist()
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num).value = header
    ws.cell(row=1, column=col_num).font = Font(bold=True)
    ws.cell(row=1, column=col_num).alignment = Alignment(horizontal='center')

# Write data (full text)
for r_idx, row in enumerate(dataframe_to_rows(table_df, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx).value = value
        ws.cell(row=r_idx, column=c_idx).alignment = Alignment(wrap_text=True, vertical='top')  # Enable text wrapping

# Apply conditional formatting
score_col = table_df.columns.get_loc('Score') + 1  # Column index for Score
presence_col = table_df.columns.get_loc('Presence') + 1  # Column index for Presence

# Score formatting: red (0), yellow (1), green (2)
red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
yellow_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
green_fill = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=score_col, max_col=score_col):
    cell = row[0]
    if cell.value == 0:
        cell.fill = red_fill
    elif cell.value == 1:
        cell.fill = yellow_fill
    elif cell.value == 2:
        cell.fill = green_fill

# Presence formatting: red (0), green (1)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=presence_col, max_col=presence_col):
    cell = row[0]
    if cell.value == 0:
        cell.fill = red_fill
    elif cell.value == 1:
        cell.fill = green_fill

# Adjust column widths based on content
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min((max_length + 2) * 1.2, 100)  # Cap width at 100 for readability
    ws.column_dimensions[column].width = adjusted_width

# Enable auto-filter for headers
ws.auto_filter.ref = ws.dimensions

# Save Excel file
excel_output_path = '/Users/nfarzi/Documents/nextpoint/results/evaluation/Anenen_Steve_evaluation.xlsx'
wb.save(excel_output_path)
print(f"Saved Excel file to {excel_output_path}")

# Set up plotting style
plt.style.use('ggplot')
sns.set_palette('colorblind')

# 1. Bar Chart for Consolidated Averages
plt.figure(figsize=(12, 6))
bars = plt.bar(avg_df.index, avg_df['Average Score'], color='skyblue', edgecolor='navy')
plt.xlabel('Consolidated ID', fontsize=12)
plt.ylabel('Average Score', fontsize=12)
plt.title('Average Scores for Consolidated Nuggets', fontsize=14)
plt.ylim(0, 2)
plt.xticks(rotation=45, ha='right')

# Add value labels on top of bars
for bar in bars:
    height = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2, height, f'{height:.2f}', 
             ha='center', va='bottom', fontsize=10)

plt.tight_layout()
plt.savefig('consolidated_averages_bar_chart.png', dpi=300)
plt.close()

# 2. Stacked Bar Chart for Score Distribution
plt.figure(figsize=(12, 6))
score_df.plot(kind='bar', stacked=True, figsize=(12, 6), 
              color=['#FF6B6B', '#FFD93D', '#6BCB77'], edgecolor='black')
plt.xlabel('Consolidated ID', fontsize=12)
plt.ylabel('Number of Nuggets', fontsize=12)
plt.title('Score Distribution by Consolidated Nugget', fontsize=14)
plt.xticks(rotation=45, ha='right')
plt.legend(title='Score', loc='upper right')
plt.tight_layout()
plt.savefig('score_distribution_stacked_bar_chart.png', dpi=300)
plt.close()

# 3. Table with Conditional Formatting (Console Output)
def color_format(val, is_score=False, is_presence=False):
    if is_score:
        if val == 0:
            return colored(val, 'red')
        elif val == 1:
            return colored(val, 'yellow')
        elif val == 2:
            return colored(val, 'green')
    if is_presence:
        if val == 0:
            return colored(val, 'red')
        elif val == 1:
            return colored(val, 'green')
    return val

# Prepare console table with truncated text for readability
console_table = table_df.copy()
console_table['Consolidated Nugget Text'] = console_table['Consolidated Nugget Text'].apply(
    lambda x: x[:50] + '...' if isinstance(x, str) and len(x) > 50 else x
)
console_table['Text'] = console_table['Text'].apply(
    lambda x: x[:50] + '...' if isinstance(x, str) and len(x) > 50 else x
)
console_table['Explanation'] = console_table['Explanation'].apply(
    lambda x: x[:50] + '...' if isinstance(x, str) and len(x) > 50 else x
)
console_table['Score'] = console_table['Score'].apply(lambda x: color_format(x, is_score=True))
console_table['Presence'] = console_table['Presence'].apply(lambda x: color_format(x, is_presence=True))

# Print table using tabulate
print("\nDetailed Nugget Evaluation Table:")
print(tabulate(console_table, headers='keys', tablefmt='grid', showindex=False))